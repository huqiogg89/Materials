import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def convert_csv_to_excel():
    csv_file = 'Таблица_ЦШП.csv'
    excel_file = 'Таблица_ЦШП.xlsx'
    
    # Конвертируем csv в excel через pandas
    df = pd.read_csv(csv_file, encoding='utf-8')
    df.to_excel(excel_file, index=False)
    
    # Открываем созданный файл через openpyxl для стилизации
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Задаем ширину колонок (буквы от A до G для 7 колонок)
    column_widths = {
        'A': 8,   # Номер
        'B': 8,   # Год
        'C': 15,  # Страна
        'D': 25,  # Тип технологии
        'E': 35,  # Компания
        'F': 80,  # Суть проекта
        'G': 15   # Источник
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Стили
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Светло-синий
    header_font = Font(bold=True, size=11, color="000000")
    border_style = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_text_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Применяем стили
    for row in ws.iter_rows():
        is_header = (row[0].row == 1)
        for cell in row:
            cell.border = border_style
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                # Центрируем номера и года
                if cell.column_letter in ['A', 'B', 'C', 'D', 'G']:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                else:
                    # Остальные влево с переносом
                    cell.alignment = wrap_text_alignment

    # Сохраняем красоту
    wb.save(excel_file)
    print("Success! File Таблица_ЦШП.xlsx styled successfully.")

if __name__ == "__main__":
    convert_csv_to_excel()
